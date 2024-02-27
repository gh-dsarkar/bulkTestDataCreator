import this
import time
import traceback

import openpyxl as openpyxl
from IPython.core import payload
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

import requests
import json
import xml.etree.ElementTree as ET

wb = openpyxl.load_workbook('./testData/RefRequestIDs.xlsx')
sheet = wb['Sheet1']
with open('config.json', 'r') as f:
    config = json.load(f)

# Get Credentials from COnfig

service = Service()

# Login to LIMS

options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)




# def get_BillingToken():
#     API_USERNAME = config['Billing_cred']['username']
#     API_PASSWORD = config['Billing_cred']['password']
#     endPoint = '/api/v1.0/auth/authenticate'
#     token_url = config['Billing_cred']['url'] + endPoint
#     payload = json.dumps({
#         "username": API_USERNAME,
#         "password": API_PASSWORD
#     })
#     headers = {
#         'Content-Type': 'application/json'
#     }
#
#     response = requests.request("POST", token_url, headers=headers, data=payload)
#     print('token generated successfully')
#     return response.json()['token']


# def trigger_medicarePoller(requestID, token):
#     # medicareEndPoint = "/api/v2.0/guardanthealth/billing/update?ids=" + requestID + "&types=Clinical&field=medcoverage&token=" + token
#     medicareEndPoint = f"/api/v2.0/guardanthealth/billing/update?ids=%s&types=Clinical&field=medcoverage&token=%s"%(requestID,token)
#
#     medicareUrl = config['Billing_cred']['billingurl'] + medicareEndPoint
#     payload = ""
#     headers = {
#         'Authorization': f'Bearer {token}'
#         # 'Authorization': 'Bearer ' + token
#     }
#     response = requests.request("PUT", medicareUrl, headers=headers, data=payload)
#     print('medicarePoller executed')
#     return response.status_code


# def trigger_txStatusPoller(requestID, token):
#     txStatusEndPoint = "/api/v2.0/guardanthealth/billing/update?ids=" + requestID + "&types=Clinical&token=" + token + "&field=txstatus"
#
#     txStatusUrl = config['Billing_cred']['billingurl'] + txStatusEndPoint
#     payload = ""
#     headers = {
#         'Authorization': 'Bearer ' + token
#     }
#     response = requests.request("PUT", txStatusUrl, headers=headers,data=payload)
#     print('txStatusUrl executed')
#     return response.status_code


def runActionBlockinLIMS():
    uname = config['lims']['username']
    pswrd = config['lims']['password']
    lims_url = config['lims']['url']
    action_name = "Test_AccessionRequest_withNewPatient_ACS"

    # 0login_url = "https://lims-val.ghdna.io/ghlims/logon.jsp?sso=n"
    driver.get(lims_url)
    # driver.manage().window().maximize();
    lims_username = driver.find_element("id", "databaseusername")
    lims_pwrd = driver.find_element("id", "databasepassword")
    lims_username.send_keys(uname)
    lims_pwrd.send_keys(pswrd)
    driver.find_element("id", "submitlogin").click()
    time.sleep(5)

    # Select User Role - System Admin
    # driver.find_element("id","jobtypeselector")

    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='jobtypeselector']"))).click()
    jobroleselect = Select(
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='jobtypeselector']"))))
    if (jobroleselect.first_selected_option.text != 'System Admin'):
        jobroleselect.select_by_visible_text('System Admin')

        # print([o.text for o in jobroleselect.options])
        # Check Selected Option
        driver.find_element("xpath", "//input[@type='checkbox'][@id='clearoldconnection']").click()
        SelectedOption = Select(
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//select[@id='jobtypeid']"))))
        if (SelectedOption.first_selected_option.text == 'System Admin'):
            print("System Admin Selected")
        else:
            SelectedOption.select_by_visible_text('System Admin')
        driver.find_element("xpath", "//button[@id='ok']").click()

    time.sleep(5)
    driver.find_element("xpath", "//div[@title='Menu and Dashboard Picker']").click()
    driver.find_element("xpath",
                        "//div[@class='menugizmo_div']/ul/li[5]/a[contains(text(),'System Admin Menu')]").click()

    time.sleep(5)
    driver.switch_to.frame("_nav_frame1")

    driver.find_element("xpath", "//table[@id='contenttable1']//td[contains(text(),'Actions')]").click()

    time.sleep(5)
    driver.switch_to.default_content()
    driver.switch_to.frame("_nav_frame1")
    driver.find_element("xpath", "//input[@id='listtop_basicsearchbox']").clear()
    driver.find_element("xpath", "//input[@id='listtop_basicsearchbox']").send_keys(action_name)
    driver.find_element("xpath", "//div[@title='Click to Search']").click()

    time.sleep(5)
    driver.switch_to.frame("list_iframe")
    driver.find_element("xpath", "//td/a[contains(text(),'Test_AccessionRequest_withNewPatient_ACS')]").click()

    driver.switch_to.default_content()
    time.sleep(5)

    driver.switch_to.frame("_nav_frame1")
    driver.switch_to.frame("maint_iframe")
    driver.find_element(By.XPATH,"//table//div[contains(text(),'Flow Chart')]").click()
    # driver.find_element("xpath", "//table//div[contains(text(),'Flow Chart')]").click()
    driver.find_element("xpath", "//td[contains(text(),'Edit/Test')]").click()

    driver.switch_to.default_content()
    dlg_frame0 = driver.find_element("xpath", "//*[@id='dlg_frame0']")
    driver.switch_to.frame(dlg_frame0);

    driver.find_element("xpath", "//*[@id='totest']//div[contains(text(),'Test')]").click()
    driver.switch_to.frame("rightframe")
    requestIDs,copies = getRefRequestIDs()
    # token = get_BillingToken()
    for i in range(2, len(requestIDs)):
        requestID = (requestIDs[i].value)
        noofcopies = copies[i].value
        newRequestId = performOperationInsideActionBlock(i,requestID,noofcopies)
        if newRequestId != None:
            WritNewRequestIDtoExcel(i, newRequestId)
            # medicare_resp_statusCode = trigger_medicarePoller(newRequestId, token)
            # if medicare_resp_statusCode == 200:
            #     txStatus_resp_statusCode = trigger_txStatusPoller(newRequestId, token)
            #     if txStatus_resp_statusCode == 200:
            #         time.sleep(5)
            #         getFinanceInfoResultSet(i,newRequestId)
    teardown()


def performOperationInsideActionBlock(i,requestID,noofcopies):
    copies = driver.find_element("xpath", "//input[@name='copies']")
    request = driver.find_element("xpath", "//input[@name='templateid']")
    sampleid = driver.find_element("xpath", "//input[@name='sampleid']")
    copies.clear()
    copies.send_keys(noofcopies)
    request.clear()
    request.send_keys(requestID)
    sampleID = requestID + '01'
    sampleid.clear()
    sampleid.send_keys(sampleID)
    driver.switch_to.default_content()
    dlg_frame0 = driver.find_element("xpath", "//*[@id='dlg_frame0']")
    driver.switch_to.frame(dlg_frame0);
    # Click Test Execute Button
    driver.find_element("xpath", "//*[@id='testexecute']//div[contains(text(),'Execute Now')]").click()
    driver.switch_to.frame("rightframe")
    time.sleep(5)

    # Get the new Request ID
    result = driver.find_element("xpath", "//table[2]//td[2]//tr[2]//textarea")
    time.sleep(10)
    resultStr = (driver.execute_script("return arguments[0].{0};".format('value'), result))
    # print(str(resultStr))
    if (('Untrapped Error' in resultStr) or ('GENERAL_ERROR' in resultStr) or (resultStr=='')):
        print('Error check input coming from performOperationInsideActionBlock')
        newRequestId = None
        sheet.cell(row=i, column=20).value = resultStr
        wb.save('../testData/G360_Clinical.xlsx')
    else:
        # derive new Values
        newRequestId = parseString(resultStr)
    return newRequestId


def parseString(resultStr):
    try:
        index_of_action_block = resultStr.rfind('<action')
        resultStr = resultStr[index_of_action_block: len(resultStr)]
        finalstr = resultStr[:resultStr.find('</actionblock>')]
        # print(finalstr)
        time.sleep(3)
        root = ET.fromstring(finalstr.strip())
        tag_id_to_find = "keyid1"
        xpath_expression = f".//property[@id='{tag_id_to_find}']"
        element = root.find(xpath_expression)
        if element is not None:
            # Get the value of the element
            element_value = element.text
            print(f"Value of element with ID {tag_id_to_find}: {element_value}")
            return element_value
        else:
            print(f"Element with ID {tag_id_to_find} not found in the XML.")
            return None
    except:
        print('-------FAILURE--------')
        print(finalstr)
        traceback.print_exc()



def getRefRequestIDs():
    # wb = openpyxl.load_workbook('../testData/G360_Clinical.xlsx')
    # sheet = wb['Sheet1']
    print('Total number of rows: ' + str(sheet.max_row) + '. And total number of columns: ' + str(sheet.max_column))
    requestIDs = sheet["B"]
    copies = sheet["C"]
    return requestIDs,copies


def WritNewRequestIDtoExcel(i, newRequestId):
    # wb = openpyxl.load_workbook('../testData/G360_Clinical.xlsx')
    # sheet = wb['Sheet1']
    print('Row '+str(i)+ ':+ ' +(newRequestId[:-2]))
    sheet.cell(row=i, column=5).value = newRequestId[:-2]
    wb.save('./testData/RefRequestIDs.xlsx')


# def getFinanceInfoResultSet(i,requestID):
#     resultset = LIMSDBConnect.getDatafromLIMS(requestID)
#     colnum = 9
#     for each_data in resultset:
#         sheet.cell(row=i, column=colnum).value = each_data
#         colnum=colnum+1
#     wb.save('../testData/G360_Clinical.xlsx')



def teardown():
    # LIMSDBConnect.closeConnection()
    wb.close()
    driver.close()
    driver.quit()



# requestID = 'A1075777'
# token = get_BillingToken()
# medicare_resp_statusCode = trigger_medicarePoller(requestID,token)
# if medicare_resp_statusCode == 200:
#   txStatus_resp_statusCode = trigger_txStatusPoller(requestID,token)
#   if txStatus_resp_statusCode == 200:
#     time.sleep(5)



runActionBlockinLIMS()
